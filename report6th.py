from __future__ import annotations

import json
import os
import subprocess
from datetime import date, datetime, time, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Dict, List, Tuple

import pyodbc
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


# =========================
# CONFIG
# =========================
DSN = os.getenv("POS_DSN", "pos")
UID = os.getenv("POS_UID", "db")
PWD = os.getenv("POS_PWD", "db")

# POS business-day cutoff (06:30 -> next day 06:30)
BUSINESS_DAY_START = os.getenv("BUSINESS_DAY_START", "06:30")  # "HH:MM"

# Backfill range: starting December last year (inclusive)
TODAY = date.today()
START_DATE = date(TODAY.year - 1, 12, 1)
END_DATE = TODAY  # inclusive

# Git options
DO_GIT_PUSH = True            # set False if you want to generate files only
GIT_COMMIT_MESSAGE = "Backfill reports (HTML + XLSX)"

# Repo paths (this script must sit in the repo root)
REPO_ROOT = Path(__file__).resolve().parent
REPORTS_DIR = REPO_ROOT / "reports"
INDEX_JSON = REPO_ROOT / "report_index.json"


# =========================
# Helpers
# =========================
def d0(v) -> Decimal:
    try:
        if v is None:
            return Decimal("0")
        if isinstance(v, Decimal):
            return v
        return Decimal(str(v))
    except (InvalidOperation, ValueError):
        return Decimal("0")

def fmt2(v) -> str:
    return f"{d0(v):,.2f}"

def hhmm_to_time(hhmm: str) -> time:
    hh, mm = map(int, hhmm.split(":"))
    return time(hh, mm)

def business_window(day: date) -> Tuple[datetime, datetime]:
    start = datetime.combine(day, hhmm_to_time(BUSINESS_DAY_START))
    end = start + timedelta(days=1)
    return start, end

def db_conn() -> pyodbc.Connection:
    return pyodbc.connect(f"DSN={DSN};UID={UID};PWD={PWD};", autocommit=True)

def ensure_dir(p: Path) -> None:
    p.mkdir(parents=True, exist_ok=True)

def write_text(p: Path, s: str) -> None:
    ensure_dir(p.parent)
    p.write_text(s, encoding="utf-8")

def write_xlsx(filepath: Path, sheet_name: str, headers: List[str], rows: List[List]):
    ensure_dir(filepath.parent)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    ws.append(headers)
    for r in rows:
        ws.append(r)

    # autosize columns (safe cap)
    for col in range(1, len(headers) + 1):
        max_len = 0
        for cell in ws[get_column_letter(col)]:
            v = "" if cell.value is None else str(cell.value)
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 60)

    wb.save(filepath)

def run_cmd(cmd: List[str], cwd: Path) -> subprocess.CompletedProcess:
    return subprocess.run(cmd, cwd=str(cwd), capture_output=True, text=True)

def safe_git(cmd: List[str]) -> None:
    r = run_cmd(cmd, cwd=REPO_ROOT)
    if r.returncode != 0:
        raise RuntimeError(f"Command failed: {' '.join(cmd)}\nSTDOUT:\n{r.stdout}\nSTDERR:\n{r.stderr}")


# =========================
# HTML templating
# =========================
def html_shell(title: str, body_html: str, subtitle: str = "") -> str:
    return f"""<!doctype html>
<html>
<head>
  <meta charset="utf-8">
  <title>{title}</title>
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <style>
    body {{ font-family: Arial, sans-serif; background:#0b1220; color:#e9eefc; margin:0; padding:24px; }}
    .wrap {{ max-width: 1100px; margin: 0 auto; }}
    .card {{ background:#121a2b; border:1px solid rgba(255,255,255,.08); border-radius:12px; padding:16px; }}
    .muted {{ color: rgba(233,238,252,.75); }}
    .mono {{ font-family: ui-monospace, SFMono-Regular, Menlo, Monaco, Consolas, "Liberation Mono", monospace; }}
    table {{ width:100%; border-collapse: collapse; }}
    th, td {{ padding:8px 10px; border-bottom:1px solid rgba(255,255,255,.08); vertical-align:top; }}
    thead th {{ background: rgba(255,255,255,.06); text-align:left; }}
    .right {{ text-align:right; }}
    .red {{ color:#ff6b6b; font-weight:bold; }}
  </style>
</head>
<body>
  <div class="wrap">
    <div class="card">
      <div style="display:flex; justify-content:space-between; align-items:flex-start; gap:12px;">
        <div>
          <h2 style="margin:0;">{title}</h2>
          <div class="muted">{subtitle}</div>
        </div>
        <div class="muted mono">{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</div>
      </div>
      <div style="margin-top:14px;">
        {body_html}
      </div>
    </div>
  </div>
</body>
</html>
"""

def html_table(headers: List[str], rows_html_cells: List[List[str]]) -> str:
    th = "".join(f"<th>{h}</th>" for h in headers)
    tr = ""
    for r in rows_html_cells:
        tds = "".join(f"<td>{c}</td>" for c in r)
        tr += f"<tr>{tds}</tr>"
    if not tr:
        tr = f"<tr><td colspan='{len(headers)}' class='muted'>No rows</td></tr>"
    return f"<table><thead><tr>{th}</tr></thead><tbody>{tr}</tbody></table>"


# =========================
# Report builders (POS-matching math)
# =========================
def build_daily_summary(conn: pyodbc.Connection, start_dt: datetime, end_dt: datetime, day: date) -> Tuple[str, List[List]]:
    """
    POS-matching Daily Summary core:
    - Sales base = TransType=101 only
    - Gross Total = Sum(Amount + TaxTotal*(1-TaxInclude))
    - Taxes = Sum(TaxXAmount*(1-TaxInclude))
    - Net = Gross - Taxes
    - Discount = Sum(DiscountAmount) for TransType=101
    - Customers = Count(distinct ReceiptN) where GroupTransType=1
    """
    cur = conn.cursor()
    sql = """
    SELECT
      TransType, GroupTransType, ReceiptN, SubCategoryID,
      Amount, DiscountAmount, TaxInclude,
      Tax1Amount, Tax2Amount, Tax3Amount, Tax4Amount
    FROM Journal
    WHERE DateR >= ? AND DateR < ?
      AND Status = 0
      AND (TransType IN (101,102,103,104,311,501,780) OR GroupTransType IN (1,2))
    """
    rows = cur.execute(sql, start_dt, end_dt).fetchall()
    sales = [r for r in rows if int(r.TransType or 0) == 101]

    def tax_total(r) -> Decimal:
        return d0(r.Tax1Amount) + d0(r.Tax2Amount) + d0(r.Tax3Amount) + d0(r.Tax4Amount)

    def one_minus_taxinclude(r) -> Decimal:
        return Decimal("1") - d0(r.TaxInclude)

    gross_total = sum(d0(r.Amount) + tax_total(r) * one_minus_taxinclude(r) for r in sales)
    gst = sum(d0(r.Tax1Amount) * one_minus_taxinclude(r) for r in sales)
    pst = sum(d0(r.Tax2Amount) * one_minus_taxinclude(r) for r in sales)
    liq = sum(d0(r.Tax3Amount) * one_minus_taxinclude(r) for r in sales)
    tax4 = sum(d0(r.Tax4Amount) * one_minus_taxinclude(r) for r in sales)

    total_taxes = gst + pst + liq + tax4
    net_total = gross_total - total_taxes
    discount = sum(d0(r.DiscountAmount) for r in sales)

    customers = len({int(r.ReceiptN) for r in rows if int(r.GroupTransType or 0) == 1 and r.ReceiptN is not None})
    avg_sale = (net_total / Decimal(customers)) if customers else Decimal("0")

    # Breakdown by SubCategoryID (net style)
    by_subcat: Dict[str, Decimal] = {}
    for r in sales:
        key = (str(r.SubCategoryID or "").strip() or "UNSPECIFIED")
        net_line = d0(r.Amount) - tax_total(r) * d0(r.TaxInclude)
        by_subcat[key] = by_subcat.get(key, Decimal("0")) + net_line

    # HTML table rows
    label_col = day.strftime("%B %d")
    html_rows: List[List[str]] = []
    html_rows.append([
        "<span class='red'>Total Sales:</span>",
        f"<span class='right red'>{fmt2(gross_total)}</span>",
        f"<span class='right red'>{fmt2(gross_total)}</span>",
    ])

    for k in sorted(by_subcat.keys()):
        v = by_subcat[k]
        html_rows.append([
            k,
            f"<span class='right mono'>{fmt2(v)}</span>",
            f"<span class='right mono'>{fmt2(v)}</span>",
        ])

    def add_line(label: str, val: Decimal, red: bool = False):
        cls = "red" if red else ""
        html_rows.append([
            f"<span class='{cls}'>{label}</span>",
            f"<span class='right mono {cls}'>{fmt2(val)}</span>",
            f"<span class='right mono {cls}'>{fmt2(val)}</span>",
        ])

    add_line("Net Total Sales", net_total, red=True)
    add_line("GST 5%", gst)
    add_line("PST 7%", pst)
    add_line("LIQ TAX 10%", liq)
    if tax4 != 0:
        add_line("Tax4", tax4)
    add_line("Total taxes", total_taxes, red=True)
    add_line("Total Sales", gross_total, red=True)
    add_line("Discount", discount)
    html_rows.append([
        "Customer count",
        f"<span class='right mono'>{customers}</span>",
        f"<span class='right mono'>{customers}</span>",
    ])
    add_line("Average Sale", avg_sale)

    body = html_table(["Description", label_col, "Total"], html_rows)
    subtitle = f"Business window: {start_dt} → {end_dt} | DSN={DSN}"
    html = html_shell("Summary Report Daily", body, subtitle)

    # XLSX rows (simple 2-column)
    xlsx_rows: List[List] = []
    xlsx_rows.append(["Total Sales:", float(gross_total)])
    for k in sorted(by_subcat.keys()):
        xlsx_rows.append([k, float(by_subcat[k])])
    xlsx_rows.append(["Net Total Sales", float(net_total)])
    xlsx_rows.append(["GST 5%", float(gst)])
    xlsx_rows.append(["PST 7%", float(pst)])
    xlsx_rows.append(["LIQ TAX 10%", float(liq)])
    if tax4 != 0:
        xlsx_rows.append(["Tax4", float(tax4)])
    xlsx_rows.append(["Total taxes", float(total_taxes)])
    xlsx_rows.append(["Total Sales", float(gross_total)])
    xlsx_rows.append(["Discount", float(discount)])
    xlsx_rows.append(["Customer count", int(customers)])
    xlsx_rows.append(["Average Sale", float(avg_sale)])

    return html, xlsx_rows


def build_category_report(conn: pyodbc.Connection, start_dt: datetime, end_dt: datetime, day: date) -> Tuple[str, List[List]]:
    """
    POS Category Report shape:
    Group | Amount | Amount (Taxes Included) | Category Count | Customers
    Uses Category join (C.SubCategoryID) to avoid UNSPECIFIED from Journal.SubCategoryID.
    Note: uses DateR <= end (inclusive) like POS logs.
    """
    cur = conn.cursor()

    sql = """
    SELECT
      C.SubCategoryID AS GroupName,
      SUM((J.Amount) - (J.Tax1Amount+J.Tax2Amount+J.Tax3Amount+J.Tax4Amount) * (J.TaxInclude)) AS AmountNet,
      SUM((J.Amount) + (J.Tax1Amount+J.Tax2Amount+J.Tax3Amount+J.Tax4Amount) * (1 - J.TaxInclude)) AS AmountTaxIncl,
      SUM(J.Quantity) AS CategoryCount,
      COUNT(DISTINCT J.ReceiptN) AS Customers
    FROM Journal J
      LEFT OUTER JOIN Category C ON C.CategoryID = J.CategoryID
    WHERE
      J.DateR >= ? AND J.DateR <= ?
      AND J.Status = 0
      AND (1 - C.SalesFlag) = 1
      AND J.TransType IN (101,102,112,111)
    GROUP BY C.SubCategoryID
    ORDER BY C.SubCategoryID
    """
    rows = cur.execute(sql, start_dt, end_dt).fetchall()

    sql_total = """
    SELECT
      SUM((J.Amount) - (J.Tax1Amount+J.Tax2Amount+J.Tax3Amount+J.Tax4Amount) * (J.TaxInclude)) AS AmountNet,
      SUM((J.Amount) + (J.Tax1Amount+J.Tax2Amount+J.Tax3Amount+J.Tax4Amount) * (1 - J.TaxInclude)) AS AmountTaxIncl,
      SUM(J.Quantity) AS CategoryCount,
      COUNT(DISTINCT J.ReceiptN) AS Customers
    FROM Journal J
      LEFT OUTER JOIN Category C ON C.CategoryID = J.CategoryID
    WHERE
      J.DateR >= ? AND J.DateR <= ?
      AND J.Status = 0
      AND (1 - C.SalesFlag) = 1
      AND J.TransType IN (101,102,112,111)
    """
    tot = cur.execute(sql_total, start_dt, end_dt).fetchone()

    # HTML table
    html_rows: List[List[str]] = []
    for gname, amt_net, amt_incl, qty, cust in rows:
        label = (str(gname).strip() if gname is not None else "UNSPECIFIED")
        html_rows.append([
            label,
            f"<span class='right mono'>{fmt2(amt_net)}</span>",
            f"<span class='right mono'>{fmt2(amt_incl)}</span>",
            f"<span class='right mono'>{int(qty or 0)}</span>",
            f"<span class='right mono'>{int(cust or 0)}</span>",
        ])

    html_rows.append([
        "<span class='red'>TOTAL</span>",
        f"<span class='right mono red'>{fmt2(tot.AmountNet)}</span>",
        f"<span class='right mono red'>{fmt2(tot.AmountTaxIncl)}</span>",
        f"<span class='right mono red'>{int(tot.CategoryCount or 0)}</span>",
        f"<span class='right mono red'>{int(tot.Customers or 0)}</span>",
    ])

    body = html_table(
        ["Group", "Amount", "Amount (Taxes Included)", "Category Count", "Customers"],
        html_rows
    )
    subtitle = f"Business window: {start_dt} → {end_dt} | DSN={DSN}"
    html = html_shell("Category Report", body, subtitle)

    # XLSX rows
    xlsx_rows: List[List] = []
    for gname, amt_net, amt_incl, qty, cust in rows:
        label = (str(gname).strip() if gname is not None else "UNSPECIFIED")
        xlsx_rows.append([
            label,
            float(d0(amt_net)),
            float(d0(amt_incl)),
            int(qty or 0),
            int(cust or 0),
        ])

    xlsx_rows.append([
        "TOTAL",
        float(d0(tot.AmountNet)),
        float(d0(tot.AmountTaxIncl)),
        int(tot.CategoryCount or 0),
        int(tot.Customers or 0),
    ])

    return html, xlsx_rows


# =========================
# report_index.json
# =========================
def load_index() -> Dict:
    if INDEX_JSON.exists():
        try:
            return json.loads(INDEX_JSON.read_text(encoding="utf-8"))
        except Exception:
            return {"latest": None, "dates": []}
    return {"latest": None, "dates": []}

def save_index(latest: str, dates: List[str]) -> None:
    payload = {
        "latest": latest,
        "dates": dates,
        "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    write_text(INDEX_JSON, json.dumps(payload, indent=2))


# =========================
# Main
# =========================
def main():
    print(f"Repo root: {REPO_ROOT}")
    print(f"Generating from {START_DATE} to {END_DATE} (inclusive)")
    print(f"Business start: {BUSINESS_DAY_START}")
    ensure_dir(REPORTS_DIR)

    idx = load_index()
    dates_set = set(idx.get("dates") or [])

    generated = 0
    failed: List[str] = []

    with db_conn() as conn:
        d = START_DATE
        while d <= END_DATE:
            ds = d.strftime("%Y-%m-%d")
            out_dir = REPORTS_DIR / ds
            ensure_dir(out_dir)

            start_dt, end_dt = business_window(d)

            try:
                daily_html, daily_xrows = build_daily_summary(conn, start_dt, end_dt, d)
                cat_html, cat_xrows = build_category_report(conn, start_dt, end_dt, d)

                # HTML
                write_text(out_dir / "summary_daily.html", daily_html)
                write_text(out_dir / "category_report.html", cat_html)

                # XLSX
                write_xlsx(
                    out_dir / "summary_daily.xlsx",
                    "Daily Summary",
                    ["Description", "Value"],
                    daily_xrows
                )
                write_xlsx(
                    out_dir / "category_report.xlsx",
                    "Category Report",
                    ["Group", "Amount", "Amount (Taxes Included)", "Category Count", "Customers"],
                    cat_xrows
                )

                dates_set.add(ds)
                generated += 1
                if generated % 10 == 0:
                    print(f"  generated {generated} days... (latest: {ds})")

            except Exception as e:
                failed.append(f"{ds}: {e}")
                print(f"  FAILED {ds}: {e}")

            d += timedelta(days=1)

    dates_sorted = sorted(dates_set, reverse=True)
    latest = dates_sorted[0] if dates_sorted else None
    if latest:
        save_index(latest, dates_sorted)

    print(f"\nDone. Generated days: {generated}")
    if failed:
        print("\nFailures (first 50):")
        for f in failed[:50]:
            print(" -", f)
        if len(failed) > 50:
            print(f" ... plus {len(failed) - 50} more")

    # Git add/commit/push
    if DO_GIT_PUSH:
        print("\nRunning git add/commit/push...")
        try:
            safe_git(["git", "add", "."])

            commit = run_cmd(["git", "commit", "-m", GIT_COMMIT_MESSAGE], cwd=REPO_ROOT)
            if commit.returncode != 0:
                msg = (commit.stdout + commit.stderr).lower()
                if "nothing to commit" in msg:
                    print("Nothing new to commit.")
                else:
                    raise RuntimeError(f"git commit failed:\n{commit.stdout}\n{commit.stderr}")

            safe_git(["git", "push", "origin", "main"])
            print("✅ Pushed to GitHub.")
        except Exception as e:
            print(f"⚠️ Git push failed: {e}")
            print("Try manually:\n  git push origin main")

    print("\nOpen:")
    print("  https://yan-vibt.github.io/pos-reports/")
    if latest:
        print(f"Latest date folder: reports/{latest}/")


if __name__ == "__main__":
    main()
